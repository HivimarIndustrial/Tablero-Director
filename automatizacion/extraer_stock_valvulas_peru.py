"""
Extrae el stock detallado del Grupo de Articulo 'Valvulas' desde la app
Qlik 'BI - STOCK AL CIERRE' y genera un Excel pensado para que Hivimar Peru
pueda revisar lo que hay disponible en Ecuador.

Origen Qlik:
  app:    BI - STOCK AL CIERRE  (4e165af0-9b67-48f5-8dd3-1efc9bca4c33)
  hoja:   DETALLE STOCK
  obj:    EvNDmks (table, 18 cols x ~1M filas totales)
  filtro: grupo_articulo = 'Valvulas'

Reglas (acordadas con Juan David):
  - Sin datos de cliente (no se incluye COD_CLIENTE).
  - Solo unidades y costo a libre utilizacion.
  - STOCK_ESPECIAL se traduce a TIPO_STOCK:
       '0' / 'NO CONSIGNA' -> 'Propio'
       'E' / 'V' / 'W'     -> 'Consigna'
  - Filas con UND_LIBRE_UTILIZACION = 0 se MANTIENEN.

Salida:
  Peru/Stock para Peru/Inventario_Valvulas_Peru.xlsx
"""
import os
import sys
import time
from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import qlik_client

# --- Configuracion fuente Qlik ---
APP_ID = '4e165af0-9b67-48f5-8dd3-1efc9bca4c33'
APP_NAME = 'BI - STOCK AL CIERRE'
OBJ_ID = 'EvNDmks'
SELECTIONS = {'grupo_articulo': ['Válvulas']}

# Indices de columnas en el HyperCube. OJO: los nombres del header son
# enganosos. La tabla Qlik usa columnas-etiqueta ("UND" / "ECS") seguidas
# de la columna numerica real:
#   0 GRUPO_ARTICULO   1 MARCA           2 NOMBRE_MATERIAL  3 CODIGO_MATERIAL
#   4 CENTRO           5 ALMACEN         6 STOCK_ESPECIAL   7 COD_CLIENTE
#   8 UND_LB         <- siempre texto "UND"  (etiqueta)
#   9 ECS_LB         <- numero real de unidades libre utilizacion
#  10 UND_LIBRE_UTILIZACION <- siempre texto "ECS" (etiqueta)
#  11 ECS_LIBRE_UTILIZACION <- numero real de costo libre utilizacion
#  12 UND_TRASLADO    13 ECS_TRASLADO    14 UND_CALIDAD    15 ECS_CALIDAD
#  16 UND_BLOQUEADO   17 ECS_BLOQUEADO
IDX = {
    'GRUPO': 0, 'MARCA': 1, 'NOMBRE': 2, 'CODIGO': 3,
    'CENTRO': 4, 'ALMACEN': 5, 'STOCK_ESPECIAL': 6,
    'UND_LIBRE': 9, 'COSTO_LIBRE': 11,
}

# --- Configuracion salida ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SGI_ROOT = os.path.dirname(PROJECT_DIR)
DEST_DIR = os.path.join(SGI_ROOT, 'Perú', 'Stock para Perú')
DEST_FILE = os.path.join(DEST_DIR, 'Inventario_Valvulas_Peru.xlsx')

# --- Notificacion por correo (SMTP M365) ---
EMAIL_TO = 'fab@hivimar.com'
EMAIL_CC = ''  # agregar copia si hace falta, separado por ','
EMAIL_SUBJECT = 'Inventario Valvulas Hivimar Ecuador - Stock a libre utilizacion'
SMTP_SERVICE = 'hivimar-tablero-smtp'
SMTP_HOST = 'smtp.office365.com'
SMTP_PORT = 587

OUTPUT_HEADERS = [
    'CODIGO_MATERIAL', 'NOMBRE_MATERIAL', 'MARCA',
    'CENTRO', 'ALMACEN', 'TIPO_STOCK',
    'UND_LIBRE_UTILIZACION', 'COSTO_LIBRE_UTILIZACION',
]


def traducir_tipo_stock(valor) -> str:
    s = str(valor or '').strip().upper()
    if s in ('0', '', 'NO CONSIGNA'):
        return 'Propio'
    return 'Consigna'


def transformar(headers_qlik: List[str], rows: List[list]) -> List[list]:
    """Selecciona columnas, traduce TIPO_STOCK, deja filas con UND=0."""
    out = []
    for r in rows:
        codigo  = r[IDX['CODIGO']]
        nombre  = r[IDX['NOMBRE']]
        marca   = r[IDX['MARCA']]
        centro  = r[IDX['CENTRO']]
        almacen = r[IDX['ALMACEN']]
        tipo    = traducir_tipo_stock(r[IDX['STOCK_ESPECIAL']])
        und     = r[IDX['UND_LIBRE']]
        costo   = r[IDX['COSTO_LIBRE']]
        out.append([codigo, nombre, marca, centro, almacen, tipo, und, costo])
    return out


def escribir_excel(rows: List[list], dest_path: str, fecha_corte: datetime) -> None:
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Valvulas'

    # Titulo + metadatos en filas 1-3
    ws['A1'] = 'Inventario Hivimar Ecuador - Grupo Valvulas (Stock a Libre Utilizacion)'
    ws['A1'].font = Font(bold=True, size=14, color='1F4E78')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(OUTPUT_HEADERS))

    ws['A2'] = f'Generado: {fecha_corte.strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(italic=True, color='606060')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(OUTPUT_HEADERS))

    ws['A3'] = (f'Filas: {len(rows)}  -  Origen: Qlik / {APP_NAME}  -  '
                f'Filtro: grupo_articulo = Valvulas')
    ws['A3'].font = Font(italic=True, color='606060')
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(OUTPUT_HEADERS))

    # Encabezado en fila 5
    header_row = 5
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(style='thin', color='B0B0B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, h in enumerate(OUTPUT_HEADERS, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    # Datos
    data_start = header_row + 1
    for r_idx, row in enumerate(rows, start=data_start):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            # Formato numerico para UND y COSTO
            if c_idx == 7:   # UND_LIBRE_UTILIZACION
                cell.number_format = '#,##0'
            elif c_idx == 8: # COSTO_LIBRE_UTILIZACION
                cell.number_format = '#,##0.00'

    # Anchos de columna (aproximados)
    widths = [16, 50, 18, 10, 12, 12, 18, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes y autofilter
    ws.freeze_panes = ws.cell(row=data_start, column=1)
    last_col = get_column_letter(len(OUTPUT_HEADERS))
    last_row = data_start + len(rows) - 1 if rows else data_start
    ws.auto_filter.ref = f'A{header_row}:{last_col}{last_row}'

    wb.save(dest_path)


def _construir_mensaje(remitente: str, dest_path: str, resumen: dict,
                        fecha: datetime):
    """Arma el MIMEMultipart con cuerpo HTML y el Excel adjunto."""
    import mimetypes
    from email.message import EmailMessage

    msg = EmailMessage()
    msg['From'] = remitente
    msg['To'] = EMAIL_TO
    if EMAIL_CC:
        msg['Cc'] = EMAIL_CC
    msg['Subject'] = f"{EMAIL_SUBJECT} - {fecha.strftime('%d/%m/%Y')}"

    cuerpo_html = f"""\
<html><body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #222;">
<p>Buenos d&iacute;as,</p>
<p>Adjunto el inventario actualizado del <b>Grupo de Art&iacute;culo V&aacute;lvulas</b>
   de Hivimar Ecuador, con el stock <b>a libre utilizaci&oacute;n</b>
   (unidades y costo).</p>
<p><b>Resumen al {fecha.strftime('%d/%m/%Y %H:%M')}:</b></p>
<ul>
  <li>Filas: {resumen.get('filas', 0):,}</li>
  <li>Stock Propio: {resumen.get('propio', 0):,} l&iacute;neas</li>
  <li>Stock Consigna: {resumen.get('consigna', 0):,} l&iacute;neas
      (potencialmente vendible &mdash; consultar)</li>
  <li>Total unidades libre util.: {resumen.get('und_total', 0):,.0f}</li>
  <li>Total costo libre util.: ${resumen.get('costo_total', 0):,.2f}</li>
</ul>
<p>El archivo se regenera autom&aacute;ticamente cada d&iacute;a a las 06:00 (hora Ecuador).</p>
<p style="color:#888;font-size:9pt;">Origen: Qlik &mdash; BI - STOCK AL CIERRE.
   Filtro: grupo_articulo = V&aacute;lvulas.</p>
</body></html>"""
    msg.set_content("Adjunto inventario de valvulas. Vea el HTML para el detalle.")
    msg.add_alternative(cuerpo_html, subtype='html')

    ctype, _ = mimetypes.guess_type(dest_path)
    if not ctype:
        ctype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    maintype, subtype = ctype.split('/', 1)
    with open(dest_path, 'rb') as f:
        msg.add_attachment(f.read(), maintype=maintype, subtype=subtype,
                           filename=os.path.basename(dest_path))
    return msg


def enviar_correo_outlook(dest_path: str, resumen: dict, fecha: datetime,
                           verbose: bool = True, display_only: bool = False) -> bool:
    """
    Envia el Excel adjunto via SMTP (Office 365).
    Lee credenciales del Credential Manager (servicio 'hivimar-tablero-smtp').

    En modo display_only escribe un .eml a salida_raw/ para que puedas
    abrirlo en Outlook y revisarlo, sin enviar nada.
    """
    import keyring, smtplib

    cred = keyring.get_credential(SMTP_SERVICE, None)
    if (not cred or not cred.username) and not display_only:
        print(f"  [correo] no hay credenciales en keyring '{SMTP_SERVICE}'.")
        print(f"  [correo] guarda con: python guardar_credenciales_smtp.py")
        return False

    remitente = cred.username if cred and cred.username else 'pendiente@hivimar.com'
    msg = _construir_mensaje(remitente, dest_path, resumen, fecha)

    if display_only:
        # Volcar el .eml para revision visual
        eml_path = os.path.join(SCRIPT_DIR, 'preview_correo_peru.eml')
        with open(eml_path, 'wb') as f:
            f.write(bytes(msg))
        print(f"  [correo] preview escrito a: {eml_path}")
        print(f"  [correo] doble clic al .eml para revisar (NO se envio)")
        return True

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(cred.username, cred.password)
            smtp.send_message(msg)
        if verbose:
            print(f"  [correo] enviado de {cred.username} a {EMAIL_TO}")
        return True
    except Exception as e:
        print(f"  [correo] ERROR SMTP: {type(e).__name__}: {e}")
        return False


def main():
    t0 = time.time()
    print(f"=== extraer_stock_valvulas_peru.py  {datetime.now().isoformat(timespec='seconds')} ===")
    print(f"App: {APP_NAME}  obj={OBJ_ID}")
    print(f"Filtro: {SELECTIONS}")
    print(f"Destino: {DEST_FILE}")
    print()

    cookie = qlik_client.login_and_get_cookie(verbose=True)

    headers, rows = qlik_client.fetch_table(
        app_id=APP_ID, obj_id=OBJ_ID,
        session_cookie=cookie,
        selections=SELECTIONS,
        verbose=True,
    )
    print(f"\n[qlik] headers: {headers}")
    print(f"[qlik] filas crudas: {len(rows)}")

    out_rows = transformar(headers, rows)
    print(f"[transform] filas finales: {len(out_rows)}")

    # Conteo rapido por TIPO_STOCK (col idx 5)
    propio = sum(1 for r in out_rows if r[5] == 'Propio')
    consigna = sum(1 for r in out_rows if r[5] == 'Consigna')
    und_total = sum((r[6] or 0) for r in out_rows if isinstance(r[6], (int, float)))
    costo_total = sum((r[7] or 0) for r in out_rows if isinstance(r[7], (int, float)))
    print(f"[transform] Propio: {propio}  Consigna: {consigna}")
    print(f"[transform] Total UND libre util: {und_total:,.0f}")
    print(f"[transform] Total Costo libre util: {costo_total:,.2f}")

    fecha = datetime.now()
    escribir_excel(out_rows, DEST_FILE, fecha)
    print(f"\n[xlsx] Escrito: {DEST_FILE}")

    # Enviar correo solo si no se pasa --no-mail.
    # Con --display-mail abre ventana de Outlook sin enviar (modo test).
    if '--no-mail' not in sys.argv:
        display_only = '--display-mail' in sys.argv
        print(f"\n[correo] {'Mostrando' if display_only else 'Enviando'} a {EMAIL_TO}...")
        resumen = {
            'filas': len(out_rows),
            'propio': propio, 'consigna': consigna,
            'und_total': und_total, 'costo_total': costo_total,
        }
        enviar_correo_outlook(DEST_FILE, resumen, fecha, display_only=display_only)
    else:
        print("\n[correo] saltado (--no-mail)")

    print(f"\nOK en {time.time()-t0:.1f}s")


if __name__ == '__main__':
    main()
