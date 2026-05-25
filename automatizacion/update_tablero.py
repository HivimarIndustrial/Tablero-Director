"""
Orquestador del pipeline Hivimar Industrial 8.0.

Pasos:
  1) Asegurar VPN (si no estamos en LAN HIVICORP).
  2) Descargar los 3 extractos Qlik (ventas, cartera, inventario).
  3) Descargar los 4 extractos Odoo (cotizaciones, oportunidades,
     actividades, visitas).
  4) Escribir a "Base Tablero Industria.xlsx":
     - raw_ventas                A..K
     - raw_cartera               A..AF
     - raw_inventario            A..D
     - raw_cotizaciones          A..S
     - fact_oportunidades        A..X
     - raw_actividadespendientes A..G
     - raw_visitas               A..L
     (Las columnas a la derecha con formulas se preservan intactas; si el
      tamano de datos crece/reduce, el orquestador copia/limpia formulas
      usando openpyxl Translator.)

Flags:
  --dry-run               No escribe el Excel, solo reporta los tamanos.
  --skip-vpn              Salta la verificacion/conexion VPN.
  --skip-qlik             Salta descarga Qlik (solo Odoo).
  --skip-odoo             Salta descarga Odoo (solo Qlik).
  --backup                Crea un backup del xlsx antes de escribir.
  --only <sheet>          Solo procesa esa hoja (ej: raw_ventas).
  --skip-html             Salta regeneracion del HTML del tablero.
  --write-excel           Escribe tambien las hojas raw_* del Excel (legacy;
                          por defecto ya NO se escriben - los datos van a
                          CSVs enriquecidos en salida_raw/).
"""
import os
import re
import subprocess
import sys
import shutil
import time
from datetime import datetime
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

import qlik_client
import odoo_client
import conectar_vpn
import enriquecer_datos
import logger_util
import notificar

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
XLSX = os.path.join(PROJECT_DIR, 'Base Tablero Industria.xlsx')
HTML_FILE = os.path.join(PROJECT_DIR, 'Hivimar_Tablero_Industrial 8.0.html')
REGENERAR_DB = os.path.join(PROJECT_DIR, 'regenerar_db.py')
UPDATE_HTML = os.path.join(PROJECT_DIR, 'update_html.py')
GEN_DIRECTOR = os.path.join(SCRIPT_DIR, 'generar_tablero_director.py')
PYTHON_EXE = sys.executable
SALIDA_RAW_DIR = os.path.join(PROJECT_DIR, 'salida_raw')  # CSVs para otros proyectos

# Diccionario de correcciones de nombres (typos en dim o en sistemas origen).
# Se aplica sobre los valores de texto en raw_* antes de escribir Excel / CSV.
CORRECCIONES_NOMBRES = {
    # Ejemplo, ya corregido por usuario en Excel pero queda como proteccion:
    'EDUAROD CHILAN': 'EDUARDO CHILAN',
}

# Configuracion por hoja:
#  - 'source': 'qlik' o 'odoo'
#  - 'source_key': key dentro del dict devuelto por fetch_all()
#  - 'n_data_cols': numero de columnas de datos (las primeras cols desde A)
#  - 'total_cols': total cols en la hoja (datos + formulas)
SHEETS_CONFIG = {
    'raw_ventas':                 {'source': 'qlik',  'source_key': 'ventas',        'n_data_cols': 11, 'total_cols': 22},
    'raw_cartera':                {'source': 'qlik',  'source_key': 'cartera',       'n_data_cols': 32, 'total_cols': 39},
    'raw_inventario':             {'source': 'qlik',  'source_key': 'inventario',    'n_data_cols': 4,  'total_cols': 8},
    'raw_cotizaciones':           {'source': 'odoo',  'source_key': 'cotizaciones',  'n_data_cols': 19, 'total_cols': 30},
    'fact_oportunidades':         {'source': 'odoo',  'source_key': 'oportunidades', 'n_data_cols': 24, 'total_cols': 56},
    'raw_actividadespendientes':  {'source': 'odoo',  'source_key': 'actividades',   'n_data_cols': 7,  'total_cols': 10},
    'raw_visitas':                {'source': 'odoo',  'source_key': 'visitas',       'n_data_cols': 12, 'total_cols': 16},
}


def _safe_print(s: str):
    """Imprime texto evitando UnicodeEncodeError en consolas Windows cp1252."""
    try:
        print(s)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, 'encoding', 'ascii') or 'ascii'
        print(s.encode(enc, errors='replace').decode(enc, errors='replace'))


def aplicar_correcciones(valor):
    """Aplica CORRECCIONES_NOMBRES a un valor si es string."""
    if isinstance(valor, str):
        return CORRECCIONES_NOMBRES.get(valor, valor)
    return valor


def exportar_csvs(qlik_data: dict, odoo_data: dict, verbose: bool = True) -> None:
    """
    Exporta los datos descargados como CSVs a salida_raw/.
    Estos archivos son consumibles por el otro proyecto que maneja
    indicadores en paralelo.
    """
    os.makedirs(SALIDA_RAW_DIR, exist_ok=True)

    def _dump_csv(path: str, headers: list, rows: list):
        import csv
        with open(path, 'w', encoding='utf-8', newline='') as f:
            w = csv.writer(f)
            w.writerow(headers)
            for row in rows:
                w.writerow([aplicar_correcciones(v) for v in row])
        if verbose:
            print(f"  CSV: {os.path.basename(path)}  ({len(rows)} filas)")

    # Qlik
    for key, d in (qlik_data or {}).items():
        _dump_csv(os.path.join(SALIDA_RAW_DIR, f'{key}.csv'),
                  d.get('headers', []), d.get('rows', []))
    # Odoo
    for key, d in (odoo_data or {}).items():
        _dump_csv(os.path.join(SALIDA_RAW_DIR, f'{key}.csv'),
                  d.get('headers', []), d.get('rows', []))

    # Timestamp
    with open(os.path.join(SALIDA_RAW_DIR, 'ultima_actualizacion.txt'),
              'w', encoding='utf-8') as f:
        f.write(datetime.now().isoformat(timespec='seconds') + '\n')


def backup_excel(src: str) -> str:
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = src.replace('.xlsx', f'.backup_{ts}.xlsx')
    shutil.copy2(src, dst)
    return dst


def backup_html(src: str) -> str:
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    dst = src.replace('.html', f'.backup_{ts}.html')
    shutil.copy2(src, dst)
    return dst


def inyectar_fecha_actualizacion(html_path: str, fecha_dt: datetime,
                                  verbose: bool = True) -> bool:
    """
    Sustituye el contenido del <span id="srcLbl">...</span> en el HTML con la
    fecha de ultima actualizacion. Este span esta en el header superior
    derecho junto al punto verde .dot.
    Devuelve True si sustituyo, False si no encontro el span.
    """
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except Exception as e:
        print(f"  ERROR leyendo HTML: {e}", file=sys.stderr)
        return False

    # Formato amigable en espanol: "Actualizado: 18/04/2026 21:02"
    label = f"Actualizado: {fecha_dt.strftime('%d/%m/%Y %H:%M')}"

    pattern = re.compile(r'(<span\s+id="srcLbl"[^>]*>)[^<]*(</span>)')
    new_content, n = pattern.subn(rf'\1{label}\2', content, count=1)

    if n == 0:
        if verbose:
            print(f"  [warn] no encontre <span id=\"srcLbl\"> en el HTML; no inyecto fecha")
        return False

    try:
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(new_content)
    except Exception as e:
        print(f"  ERROR escribiendo HTML: {e}", file=sys.stderr)
        return False
    if verbose:
        print(f"  fecha inyectada en header: '{label}'")
    return True


def regenerar_html_tablero(verbose: bool = True) -> int:
    """Invoca regenerar_db.py y update_html.py (sin modificarlos).
    Devuelve 0 si todo OK, != 0 si fallo.
    """
    # 1) backup del HTML actual
    if os.path.exists(HTML_FILE):
        b = backup_html(HTML_FILE)
        if verbose:
            print(f"  backup HTML: {os.path.basename(b)}")

    # 2) regenerar_db.py
    if verbose:
        print(f"  [1/2] regenerar_db.py ...")
    t0 = time.time()
    r = subprocess.run(
        [PYTHON_EXE, REGENERAR_DB],
        cwd=PROJECT_DIR, capture_output=True, text=True,
        encoding='utf-8', errors='replace',
    )
    if verbose:
        out = (r.stdout or '').strip()
        if out:
            for line in out.splitlines()[-5:]:
                _safe_print(f"    {line}")
    if r.returncode != 0:
        print(f"  ERROR regenerar_db.py (rc={r.returncode}):", file=sys.stderr)
        if r.stderr:
            print(r.stderr[-500:], file=sys.stderr)
        return r.returncode
    if verbose:
        print(f"    OK en {time.time()-t0:.1f}s")

    # 3) update_html.py
    if verbose:
        print(f"  [2/2] update_html.py ...")
    t0 = time.time()
    r = subprocess.run(
        [PYTHON_EXE, UPDATE_HTML],
        cwd=PROJECT_DIR, capture_output=True, text=True,
        encoding='utf-8', errors='replace',
    )
    if verbose:
        out = (r.stdout or '').strip()
        if out:
            for line in out.splitlines()[-5:]:
                _safe_print(f"    {line}")
    if r.returncode != 0:
        print(f"  ERROR update_html.py (rc={r.returncode}):", file=sys.stderr)
        if r.stderr:
            print(r.stderr[-500:], file=sys.stderr)
        return r.returncode
    if verbose:
        print(f"    OK en {time.time()-t0:.1f}s")

    # 4) Inyectar fecha de ultima actualizacion en el header del HTML
    if verbose:
        print(f"  [post] inyectando fecha de actualizacion...")
    inyectar_fecha_actualizacion(HTML_FILE, datetime.now(), verbose=verbose)

    # 5) Regenerar Tablero Director Industria (version reducida solo
    #    Gerencia, ~3.6 MB con string interning). Lee el HTML maestro
    #    recien actualizado y produce Tablero_Director_Industria.html.
    if verbose:
        print(f"  [post] regenerando Tablero Director Industria...")
    t0 = time.time()
    r = subprocess.run(
        [PYTHON_EXE, GEN_DIRECTOR],
        cwd=PROJECT_DIR, capture_output=True, text=True,
        encoding='utf-8', errors='replace',
    )
    if verbose:
        out = (r.stdout or '').strip()
        if out:
            for line in out.splitlines()[-6:]:
                _safe_print(f"    {line}")
    if r.returncode != 0:
        # No abortamos el pipeline si falla el Director: el tablero
        # principal ya quedo OK. Solo registramos el error.
        print(f"  WARN generar_tablero_director.py (rc={r.returncode}):",
              file=sys.stderr)
        if r.stderr:
            print(r.stderr[-500:], file=sys.stderr)
    elif verbose:
        print(f"    OK en {time.time()-t0:.1f}s")
    return 0


def detect_data_row_count(ws, n_data_cols: int, max_scan: int = 100_000) -> int:
    """
    Devuelve el indice de la ultima fila de DATOS actuales (incluye row 2..N).
    Una fila es de datos si AL MENOS una celda de A..n_data_cols no esta vacia.
    """
    last = 1  # row 1 es header
    for r in range(2, min(ws.max_row, max_scan) + 1):
        non_empty = False
        for c in range(1, n_data_cols + 1):
            if ws.cell(row=r, column=c).value not in (None, ''):
                non_empty = True
                break
        if non_empty:
            last = r
    return last


def write_sheet(ws, data: Dict, n_data_cols: int, total_cols: int,
                verbose: bool = True) -> dict:
    """
    Escribe las filas de `data['rows']` a las columnas 1..n_data_cols.
    Maneja el crecimiento/reduccion copiando o limpiando formulas
    en columnas n_data_cols+1..total_cols.
    Devuelve dict con stats.
    """
    new_rows = data['rows']
    n_new = len(new_rows)

    # Filas de datos existentes (antes de escribir)
    old_last = detect_data_row_count(ws, n_data_cols)
    old_data = max(0, old_last - 1)

    # Nuevo last row
    new_last = 1 + n_new

    if verbose:
        print(f"  existentes: {old_data} filas de datos (hasta row {old_last})")
        print(f"  nuevas    : {n_new} filas   (hasta row {new_last})")

    # 1) Capturar formulas template de la fila 2 (para propagar si crecemos)
    formula_templates = {}
    if total_cols > n_data_cols:
        for c in range(n_data_cols + 1, total_cols + 1):
            tpl_cell = ws.cell(row=2, column=c)
            v = tpl_cell.value
            if isinstance(v, str) and v.startswith('='):
                formula_templates[c] = v

    # 2) Escribir filas nuevas (columnas 1..n_data_cols)
    for i, row_values in enumerate(new_rows):
        r = 2 + i
        for j in range(n_data_cols):
            v = row_values[j] if j < len(row_values) else None
            ws.cell(row=r, column=j + 1).value = v

    # 3) Crecimiento: propagar formulas a las nuevas filas extra (si hay)
    if new_last > old_last and formula_templates:
        for r in range(old_last + 1, new_last + 1):
            for c, tpl in formula_templates.items():
                # Translator: ajusta referencias relativas al nuevo row
                try:
                    translated = Translator(tpl, origin=f"{get_column_letter(c)}2") \
                                 .translate_formula(f"{get_column_letter(c)}{r}")
                except Exception:
                    translated = tpl
                ws.cell(row=r, column=c).value = translated

    # 4) Reduccion: limpiar filas sobrantes (A..total_cols) de new_last+1 a old_last
    if new_last < old_last:
        for r in range(new_last + 1, old_last + 1):
            for c in range(1, total_cols + 1):
                ws.cell(row=r, column=c).value = None

    return {
        'old_data_rows': old_data,
        'new_data_rows': n_new,
        'formulas_propagated': max(0, new_last - old_last) * len(formula_templates),
        'rows_cleared': max(0, old_last - new_last),
    }


def main():
    args = sys.argv[1:]
    dry_run = '--dry-run' in args
    skip_vpn = '--skip-vpn' in args
    skip_qlik = '--skip-qlik' in args
    skip_odoo = '--skip-odoo' in args
    skip_html = '--skip-html' in args
    html_only = '--html-only' in args  # solo regenerar el HTML, no tocar Excel
    backup = '--backup' in args
    write_excel = '--write-excel' in args  # legacy, por defecto NO se escribe Excel raw_*
    only = None
    if '--only' in args:
        only = args[args.index('--only') + 1]

    if html_only:
        print(f"=== update_tablero.py (--html-only) ===")
        rc = regenerar_html_tablero()
        sys.exit(rc)

    t_start = time.time()
    print(f"=== update_tablero.py  {datetime.now().isoformat(timespec='seconds')} ===")
    print(f"Excel: {XLSX}")
    if dry_run:
        print("*** DRY RUN: no se escribira el Excel ***")
    print()

    # 1) VPN
    if not skip_vpn:
        print("--- Paso 1: VPN ---")
        r = conectar_vpn.ensure_vpn()
        print(f"  VPN: {r['action']}  (vpn_connected={r['vpn_connected']}, "
              f"in_corp_lan={r['in_corp_lan']})")
        if not (r['vpn_connected'] or r['in_corp_lan']):
            sys.exit("FATAL: no hay conectividad a la red interna")
        print()

    # 2) Qlik
    qlik_data = {}
    rotacion_data = None
    segmento_data = None
    if not skip_qlik:
        print("--- Paso 2: Qlik ---")
        qlik_data = qlik_client.fetch_all(verbose=True)
        print()
        # 2b) Rotacion desde BI - SKU PROFILER (ventas+stock de toda la empresa)
        print("--- Paso 2b: Qlik ROTACION (SKU Profiler) ---")
        try:
            rotacion_data = qlik_client.fetch_rotacion(verbose=True)
            print()
        except Exception as e:
            print(f"  ERROR extrayendo rotacion: {e}", file=sys.stderr)
            print(f"  continuando sin rotacion")
            rotacion_data = None
            print()
        # 2c) Segmentacion de clientes (cod_cliente -> SEGMENTO_CLIENTE detallado)
        print("--- Paso 2c: Qlik SEGMENTACION CLIENTES (Ventas) ---")
        try:
            segmento_data = qlik_client.fetch_segmentacion_clientes(verbose=True)
            print()
        except Exception as e:
            print(f"  ERROR extrayendo segmentacion: {e}", file=sys.stderr)
            print(f"  continuando sin segmentacion")
            segmento_data = None
            print()

        # 2d) Inventario por SKU (BI - STOCK AL CIERRE) -> CSV para Cotizador por Correo
        # Es independiente del tablero; si falla NO debe abortar el pipeline.
        print("--- Paso 2d: Qlik INVENTARIO POR SKU (Stock al Cierre) ---")
        try:
            import extraer_inventario_sku
            from datetime import datetime as _dt
            headers_inv, rows_inv = extraer_inventario_sku.fetch_inventario_sku(verbose=True)
            extraer_inventario_sku.escribir_csv(
                rows_inv, extraer_inventario_sku.DEST_FILE, _dt.now()
            )
            print(f"  CSV inventario SKU: {extraer_inventario_sku.DEST_FILE} "
                  f"({len(rows_inv)} SKUs)")
            print()
        except Exception as e:
            print(f"  ERROR extrayendo inventario SKU: {e}", file=sys.stderr)
            print(f"  continuando sin actualizar inventario_sku.csv")
            print()

        # 2e) Ventas Industria (historico 36 meses por (cliente, SKU))
        #     -> CSV unico para Cotizador por Correo (alimenta MBB con
        #        PVP de la ultima venta).
        # Filtro: clientes asignados a supervisores Industria
        # (JJJ/JR/DM/PS/VQ/MB) segun dim_clientes.
        print("--- Paso 2e: Qlik VENTAS INDUSTRIA (36m por cliente x SKU) ---")
        try:
            import extraer_ventas_industria
            _t = time.time()
            _rows, _ultimas = extraer_ventas_industria.fetch_ventas_industria(
                verbose=True
            )
            _dest = os.path.join(extraer_ventas_industria.DEST_DIR,
                                 extraer_ventas_industria.DEST_FILE)
            _n = extraer_ventas_industria.escribir_csv(
                _rows, _ultimas, _dest, verbose=True
            )
            print(f"  Industria: {_n:,} filas en {_dest} "
                  f"({time.time()-_t:.0f}s)")
            print()
        except Exception as e:
            print(f"  ERROR extrayendo ventas Industria: {e}", file=sys.stderr)
            print(f"  continuando sin actualizar ventas_industria.csv")
            print()

        # 2f) Digitador Industria (24 meses por digitador x fecha x factura)
        #     -> CSV en Cotizador por Correo\historico_clientes para proyecto
        #        de productividad/decisiones. Sesion Qlik independiente,
        #        envuelta en try/except: ningun fallo aqui detiene el resto
        #        del pipeline ni toca archivos de salida_raw/ (tableros).
        print("--- Paso 2f: Qlik DIGITADOR INDUSTRIA (24m granular factura) ---")
        try:
            import extraer_digitador_industria
            _t = time.time()
            _rows_d = extraer_digitador_industria.fetch_digitadores(verbose=True)
            _dest_d = os.path.join(extraer_digitador_industria.DEST_DIR,
                                   extraer_digitador_industria.DEST_FILE)
            _n_d = extraer_digitador_industria.escribir_csv(
                _rows_d, _dest_d, verbose=True
            )
            print(f"  Digitador: {_n_d:,} filas en {_dest_d} "
                  f"({time.time()-_t:.0f}s)")
            print()
        except Exception as e:
            print(f"  ERROR extrayendo digitador Industria: {e}", file=sys.stderr)
            print(f"  continuando sin actualizar digitadores_industria.csv "
                  f"(no afecta tableros)")
            print()

        # 2g) Entregas Hivitrack Industria (incremental 45d, upsert sobre
        #     CSV historico). Alimenta la pestana 'Entregas' del Tablero
        #     Director. Aislada en try/except; si falla, NO toca el
        #     entregas_industria.csv previo.
        print("--- Paso 2g: Hivitrack ENTREGAS INDUSTRIA (incremental 45d) ---")
        try:
            import extraer_entregas_hivitrack as ent_h
            _t = time.time()
            _nuevas = ent_h.descargar_ventana(verbose=True)
            _hist = ent_h.cargar_historico()
            _hist = ent_h.upsert(_hist, _nuevas, verbose=True)
            ent_h.escribir_csv(_hist, verbose=True)
            print(f"  Entregas Industria: {len(_hist):,} filas en historico "
                  f"({time.time()-_t:.0f}s)")
            print()
        except Exception as e:
            print(f"  ERROR extrayendo entregas Hivitrack: {e}", file=sys.stderr)
            print(f"  continuando sin actualizar entregas_industria.csv "
                  f"(la pestana Entregas usara la version previa)")
            print()

    # 3) Odoo
    odoo_data = {}
    if not skip_odoo:
        print("--- Paso 3: Odoo ---")
        odoo_data = odoo_client.fetch_all(verbose=True)
        print()

        # 3b) MB58 (consignment.line + check_agreement) -> mb58.csv
        print("--- Paso 3b: Odoo MB58 (consignment / prestamos) ---")
        try:
            import extraer_mb58
            _t = time.time()
            _cli = odoo_client.OdooClient(verbose=False)
            _cli.authenticate()
            _rows = extraer_mb58.extract_mb58(_cli)
            extraer_mb58.save_csv(_rows)
            _no_reg = sum(1 for r in _rows if r['es_no_regularizado'])
            print(f"  MB58: {len(_rows):,} lineas, {_no_reg:,} no regularizadas "
                  f"({time.time()-_t:.0f}s)")
            print()
        except Exception as e:
            print(f"  ERROR extrayendo MB58: {e}", file=sys.stderr)
            print(f"  continuando sin actualizar mb58.csv")
            print()

    # 4) Enriquecer datos (aplica VLOOKUPs en Python) y escribir CSVs
    print("--- Paso 4: Enriquecer datos + escribir CSVs ---")
    if dry_run:
        print("  DRY RUN: solo reportando tamanos")
        for sheet, cfg in SHEETS_CONFIG.items():
            bucket = qlik_data if cfg['source'] == 'qlik' else odoo_data
            d = bucket.get(cfg['source_key'], {})
            n = len(d.get('rows', []))
            print(f"  {sheet:30}  {n} filas nuevas")
        print()
        print(f"Tiempo total: {time.time()-t_start:.1f}s")
        return

    # Enriquecer y escribir CSVs enriquecidos (los que consume regenerar_db.py)
    # clientes_vendedor (res.partner.user_id) viene en odoo_data; lo pasamos
    # aparte para que enriquecer_datos pueble vendedor_cartera_std.
    clientes_vendedor_data = odoo_data.get('clientes_vendedor') if odoo_data else None
    enriquecer_datos.exportar_csvs(qlik_data, odoo_data, SALIDA_RAW_DIR,
                                    segmento_data=segmento_data,
                                    clientes_vendedor_data=clientes_vendedor_data,
                                    verbose=True)
    # Tambien exportar los CSVs crudos (sin enriquecimiento) para otros proyectos
    exportar_csvs(qlik_data, odoo_data)
    print(f"  CSVs enriquecidos + crudos disponibles en: {SALIDA_RAW_DIR}")

    # Rotacion: csv directo (no necesita enriquecimiento, ya viene agrupado por
    # grupo_articulo + marca + AÑO + MES con salidas y stock).
    if rotacion_data:
        import csv
        os.makedirs(SALIDA_RAW_DIR, exist_ok=True)
        rot_path = os.path.join(SALIDA_RAW_DIR, 'rotacion_profiler.csv')
        with open(rot_path, 'w', encoding='utf-8', newline='') as f:
            w = csv.writer(f)
            w.writerow(rotacion_data['headers'])
            for row in rotacion_data['rows']:
                w.writerow(row)
        print(f"  CSV rotacion: {os.path.basename(rot_path)} ({len(rotacion_data['rows'])} filas)")

    # Segmentacion clientes: csv directo (cod_cliente + SEGMENTO_CLIENTE)
    if segmento_data:
        import csv
        os.makedirs(SALIDA_RAW_DIR, exist_ok=True)
        seg_path = os.path.join(SALIDA_RAW_DIR, 'segmentacion_clientes.csv')
        with open(seg_path, 'w', encoding='utf-8', newline='') as f:
            w = csv.writer(f)
            w.writerow(segmento_data['headers'])
            for row in segmento_data['rows']:
                w.writerow(row)
        print(f"  CSV segmentacion: {os.path.basename(seg_path)} ({len(segmento_data['rows'])} filas)")

    # 4b) Escritura de Excel raw_* (legacy - solo si se pide con --write-excel)
    if write_excel:
        print()
        print("--- Paso 4b: Escribir hojas raw_* del Excel (modo legacy) ---")
        if backup:
            b = backup_excel(XLSX)
            print(f"  backup: {b}")
        wb = load_workbook(XLSX, keep_vba=False, data_only=False)
        for sheet, cfg in SHEETS_CONFIG.items():
            if only and sheet != only:
                continue
            if sheet not in wb.sheetnames:
                print(f"  [!!] hoja '{sheet}' no existe en el libro, se omite")
                continue
            bucket = qlik_data if cfg['source'] == 'qlik' else odoo_data
            if cfg['source_key'] not in bucket:
                print(f"  [SKIP] {sheet}: no se descargo '{cfg['source_key']}'")
                continue
            d = bucket[cfg['source_key']]
            print(f"\n  [HOJA] {sheet}")
            ws = wb[sheet]
            stats = write_sheet(ws, d, cfg['n_data_cols'], cfg['total_cols'])
            print(f"    stats: {stats}")
        print()
        print(f"Guardando {XLSX}...")
        wb.save(XLSX)
        print(f"Excel guardado.")
    else:
        print()
        print("  [info] Excel raw_* NO se escribio (modo nuevo). Usa --write-excel")
        print("         si necesitas que las hojas raw_* del Excel se actualicen.")

    # 5) Regenerar HTML del tablero
    if not skip_html:
        print()
        print("--- Paso 5: Regenerar HTML del tablero ---")
        rc = regenerar_html_tablero()
        if rc != 0:
            print("FATAL: fallo la regeneracion del HTML", file=sys.stderr)
            sys.exit(rc)

    # 6) Cifrar Tablero Director + push a GitHub Pages (publicacion publica)
    #    Si falla, NO abortamos: el HTML local ya quedo bien.
    #    Skipea con --skip-publicar.
    if '--skip-publicar' not in args and not skip_html:
        print()
        print("--- Paso 6: Cifrar Tablero Director + push a GitHub Pages ---")
        t_pub = time.time()
        try:
            cifrar_script = os.path.join(PROJECT_DIR, 'cifrar_y_publicar.py')
            r = subprocess.run(
                [PYTHON_EXE, cifrar_script],
                cwd=PROJECT_DIR, capture_output=True, text=True,
                encoding='utf-8', errors='replace',
            )
            if r.stdout:
                for line in r.stdout.strip().splitlines()[-5:]:
                    _safe_print(f"    {line}")
            if r.returncode != 0:
                print(f"  WARN cifrar_y_publicar.py (rc={r.returncode}):",
                      file=sys.stderr)
                if r.stderr:
                    print(r.stderr[-500:], file=sys.stderr)
            else:
                # Git: add docs/ + commit + push. Si no hay cambios, git
                # commit devuelve rc=1 inofensivo (lo tratamos como OK).
                fecha_str = datetime.now().strftime('%d/%m/%Y %H:%M')
                msg = f"Actualizacion diaria {fecha_str}"
                git_cmds = [
                    ['git', 'add', 'docs/index.html'],
                    ['git', 'commit', '-m', msg],
                    ['git', 'push'],
                ]
                git_ok = True
                for cmd in git_cmds:
                    rg = subprocess.run(cmd, cwd=PROJECT_DIR,
                                         capture_output=True, text=True,
                                         encoding='utf-8', errors='replace')
                    label = ' '.join(cmd[:2])
                    if rg.returncode != 0:
                        out = (rg.stdout or '') + (rg.stderr or '')
                        # Caso comun: "nothing to commit" -> no es error real.
                        if 'nothing to commit' in out.lower():
                            print(f"    {label}: sin cambios (docs/ ya publicado)")
                            git_ok = False  # no tiene sentido push despues
                            break
                        print(f"    {label} fallo: {out.strip()[-300:]}",
                              file=sys.stderr)
                        git_ok = False
                        break
                    else:
                        print(f"    {label}: OK")
                if git_ok:
                    print(f"    push a GitHub Pages: OK en {time.time()-t_pub:.1f}s")
        except Exception as e:
            print(f"  WARN paso publicar: {e}", file=sys.stderr)

    # 7) Stock Valvulas para Peru (extractor + correo automatico).
    #    Pipeline aparte: si falla, no abortamos el resto. Skipea con --skip-peru.
    if '--skip-peru' not in args:
        print()
        print("--- Paso 6: Stock Valvulas Peru ---")
        t_p = time.time()
        peru_script = os.path.join(SCRIPT_DIR, 'extraer_stock_valvulas_peru.py')
        r = subprocess.run(
            [PYTHON_EXE, peru_script],
            cwd=SCRIPT_DIR, capture_output=True, text=True,
            encoding='utf-8', errors='replace',
        )
        out = (r.stdout or '').strip()
        if out:
            for line in out.splitlines()[-8:]:
                _safe_print(f"    {line}")
        if r.returncode != 0:
            print(f"  WARN extraer_stock_valvulas_peru.py (rc={r.returncode}):",
                  file=sys.stderr)
            if r.stderr:
                print(r.stderr[-500:], file=sys.stderr)
        else:
            print(f"    OK en {time.time()-t_p:.1f}s")

    print()
    print(f"OK. Tiempo total: {time.time()-t_start:.1f}s")

    # Construir resumen para la notificacion de exito
    resumen = [f"Tiempo total: {time.time()-t_start:.1f}s"]
    if qlik_data:
        for k, d in qlik_data.items():
            resumen.append(f"  {k}: {len(d.get('rows', []))} filas")
    if odoo_data:
        for k, d in odoo_data.items():
            resumen.append(f"  {k}: {len(d.get('rows', []))} filas")
    if rotacion_data:
        resumen.append(f"  rotacion: {len(rotacion_data.get('rows', []))} filas")
    if segmento_data:
        resumen.append(f"  segmentacion: {len(segmento_data.get('rows', []))} filas")
    return "\n".join(resumen)


def main_with_handling():
    """Envuelve main() con logging detallado y notificacion de error."""
    # Limpiar logs antiguos (>30 dias) antes de crear uno nuevo
    try:
        logger_util.cleanup_logs(dias=30)
    except Exception:
        pass
    # Iniciar logging a archivo
    log_path, tee = logger_util.setup_logging('update_tablero')
    resumen = ""
    exit_code = 0
    try:
        resumen = main() or ""
        # Exito: notificar
        try:
            notificar.notificar_exito(resumen or "Pipeline completado sin detalles.")
        except Exception as e:
            print(f"  [warn] notificar_exito fallo: {e}")
    except SystemExit as e:
        exit_code = int(e.code) if e.code is not None else 1
        if exit_code != 0:
            # Error con SystemExit - capturar detalles
            tb = notificar.capture_traceback(e)
            try:
                notificar.notificar_error(
                    f"SystemExit({exit_code}): {e}",
                    log_path=log_path,
                    traceback_txt=tb,
                )
            except Exception as ne:
                print(f"  [warn] notificar_error fallo: {ne}")
        else:
            # SystemExit(0) = fin normal
            try:
                notificar.notificar_exito(resumen or "Pipeline completado.")
            except Exception:
                pass
    except KeyboardInterrupt:
        exit_code = 130
        print("\nINTERRUMPIDO POR USUARIO (Ctrl+C)")
        try:
            notificar.notificar_error(
                "Pipeline interrumpido por el usuario (Ctrl+C).",
                log_path=log_path,
                traceback_txt=None,
            )
        except Exception:
            pass
    except Exception as e:
        exit_code = 1
        tb = notificar.capture_traceback(e)
        print(f"\nFATAL: {type(e).__name__}: {e}")
        print(tb)
        try:
            notificar.notificar_error(
                f"{type(e).__name__}: {e}",
                log_path=log_path,
                traceback_txt=tb,
            )
        except Exception as ne:
            print(f"  [warn] notificar_error fallo: {ne}")
    finally:
        logger_util.finalize_logging(tee, exit_code=exit_code,
                                     error=None if exit_code == 0 else "Ver arriba")
        print(f"\nLog guardado en: {log_path}")
        sys.exit(exit_code)


if __name__ == '__main__':
    main_with_handling()
