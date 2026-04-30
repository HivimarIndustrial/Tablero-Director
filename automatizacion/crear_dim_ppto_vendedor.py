"""
Crea la hoja 'dim_ppto_vendedor' DIRECTAMENTE en el Excel principal,
extrayendo los datos de presupuesto limpios de fact_vtas_vs_ppto.

Estructura de la nueva hoja (5 columnas):
  vendedor_principal_std | grupo_articulo_std | anio | mes | presupuesto

Si la hoja ya existe, se sobreescribe (se borra y se vuelve a crear).

IMPORTANTE: cierra el Excel antes de correr este script.
"""
import os
import sys
import shutil
from datetime import datetime
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
XLSX = os.path.join(PROJECT_DIR, 'Base Tablero Industria.xlsx')

# Backup antes de tocar
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
BACKUP = XLSX.replace('.xlsx', f'.preppto_{ts}.xlsx')
print(f"Backup: {BACKUP}")
shutil.copy2(XLSX, BACKUP)

print(f"Leyendo: {XLSX}")
wb = load_workbook(XLSX, data_only=True)  # NO read_only (vamos a escribir)
ws = wb['fact_vtas_vs_ppto']

# Leer headers (fila 1)
header_row = [c.value for c in ws[1]]
print(f"Headers originales: {header_row}")

# Identificar indices
idx = {h: i for i, h in enumerate(header_row)}

col_vend = idx.get('vendedor_principal_std')
col_grupo = idx.get('grupo_articulo_std')
col_anio = idx.get('anio')
col_mes = idx.get('mes')
col_ppto = idx.get('presupuesto')

if None in (col_vend, col_grupo, col_anio, col_mes, col_ppto):
    print(f"ERROR: faltan columnas. idx={idx}")
    sys.exit(1)

# Recolectar filas agregando por vendedor+grupo+anio+mes
acumul = {}
total = 0
con_ppto = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    total += 1
    if not row or len(row) <= max(col_vend, col_grupo, col_anio, col_mes, col_ppto):
        continue
    vend = row[col_vend]
    grupo = row[col_grupo]
    anio = row[col_anio]
    mes = row[col_mes]
    ppto = row[col_ppto]

    if not vend or not grupo or not anio or not mes:
        continue
    try:
        ppto_num = float(ppto) if ppto is not None else 0
    except (ValueError, TypeError):
        ppto_num = 0

    if ppto_num <= 0:
        continue  # ignoramos filas sin ppto

    con_ppto += 1
    key = (str(vend).strip(), str(grupo).strip(), int(anio), int(mes))
    if key not in acumul:
        acumul[key] = 0
    acumul[key] += ppto_num

print(f"Filas originales: {total}")
print(f"Filas con ppto > 0: {con_ppto}")
print(f"Filas unicas (vendedor+grupo+anio+mes) con ppto: {len(acumul)}")

# Ordenar por anio, mes, vendedor, grupo
sorted_keys = sorted(acumul.keys(), key=lambda k: (k[2], k[3], k[0], k[1]))

# Si la hoja ya existe, la removemos para recrearla limpia
SHEET_NAME = 'dim_ppto_vendedor'
if SHEET_NAME in wb.sheetnames:
    print(f"La hoja '{SHEET_NAME}' ya existe - la recreamos")
    del wb[SHEET_NAME]

ws_out = wb.create_sheet(SHEET_NAME)
ws_out.append(['vendedor_principal_std', 'grupo_articulo_std', 'anio', 'mes', 'presupuesto'])

for vend, grupo, anio, mes in sorted_keys:
    ws_out.append([vend, grupo, anio, mes, round(acumul[(vend, grupo, anio, mes)], 2)])

# Stats por anio-mes
por_mes = {}
for (_, _, anio, mes), val in acumul.items():
    km = f"{anio}-{mes:02d}"
    por_mes[km] = por_mes.get(km, 0) + val
print()
print("Totales de ppto por anio-mes:")
for km in sorted(por_mes):
    print(f"  {km}: ${por_mes[km]:>15,.2f}")

print()
print(f"Guardando Excel con nueva hoja '{SHEET_NAME}'...")
wb.save(XLSX)
wb.close()
print(f"OK. Filas escritas: {len(acumul)}")
print(f"La hoja quedo en: {XLSX} -> '{SHEET_NAME}'")
